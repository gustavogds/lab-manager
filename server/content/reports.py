import io
import json
from zoneinfo import ZoneInfo

from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from content.models import ResearchArea, Project, Partnership, Equipment, Room, RoomSection
from accounts.models import User


UI_STRINGS = {
    "pt": {
        "active": "Ativo",
        "inactive": "Inativo",
        "generated_at": "Gerado em",
        "datetime_format": "%d/%m/%Y às %H:%M",
        "no_records": "Nenhum registro encontrado.",
        "no_room": "Sem Sala",
    },
    "en": {
        "active": "Active",
        "inactive": "Inactive",
        "generated_at": "Generated on",
        "datetime_format": "%m/%d/%Y at %H:%M",
        "no_records": "No records found.",
        "no_room": "No Room",
    },
}


def _normalize_lang(value):
    return "pt" if str(value or "").lower().startswith("pt") else "en"


def _resolve_timezone(value):
    """Resolve an IANA timezone name (e.g. from the browser), falling back to UTC."""
    if value:
        try:
            return ZoneInfo(str(value))
        except Exception:
            pass
    return ZoneInfo("UTC")


def _localized(obj, field, lang):
    """Return obj.<field>_<lang>, falling back to the other language when empty."""
    if obj is None:
        return ""
    primary = getattr(obj, f"{field}_{lang}", "") or ""
    if primary:
        return primary
    other = "en" if lang == "pt" else "pt"
    return getattr(obj, f"{field}_{other}", "") or ""


def _status_label(is_active, lang):
    return UI_STRINGS[lang]["active"] if is_active else UI_STRINGS[lang]["inactive"]


COLUMN_CONFIG = {
    "research_areas": {
        "label_pt": "Áreas de Pesquisa",
        "label_en": "Research Areas",
        "columns": [
            {"key": "id", "label_pt": "ID", "label_en": "ID", "getter": lambda item, lang: item.id},
            {"key": "title", "label_pt": "Título", "label_en": "Title", "getter": lambda item, lang: _localized(item, "title", lang)},
            {"key": "description", "label_pt": "Descrição", "label_en": "Description", "getter": lambda item, lang: _localized(item, "description", lang)},
            {"key": "status", "label_pt": "Status", "label_en": "Status", "getter": lambda item, lang: _status_label(item.is_active, lang)},
        ],
        "get_data": lambda: ResearchArea.objects.all().order_by("order", "title_pt"),
        "supports_room_grouping": False,
        "supports_section_grouping": False,
    },
    "projects": {
        "label_pt": "Projetos",
        "label_en": "Projects",
        "columns": [
            {"key": "id", "label_pt": "ID", "label_en": "ID", "getter": lambda item, lang: item.id},
            {"key": "title", "label_pt": "Título", "label_en": "Title", "getter": lambda item, lang: _localized(item, "title", lang)},
            {"key": "description", "label_pt": "Descrição", "label_en": "Description", "getter": lambda item, lang: _localized(item, "description", lang)},
            {"key": "members", "label_pt": "Integrantes", "label_en": "Members", "getter": lambda item, lang: ", ".join(m.name for m in item.members.filter(is_public=True))},
            {"key": "status", "label_pt": "Status", "label_en": "Status", "getter": lambda item, lang: _status_label(item.is_active, lang)},
        ],
        "get_data": lambda: Project.objects.prefetch_related("members").all().order_by("order", "title_pt"),
        "supports_room_grouping": False,
        "supports_section_grouping": False,
    },
    "users": {
        "label_pt": "Usuários",
        "label_en": "Users",
        "columns": [
            {"key": "id", "label_pt": "ID", "label_en": "ID", "getter": lambda item, lang: item.id},
            {"key": "name", "label_pt": "Nome", "label_en": "Name", "getter": lambda item, lang: item.name or ""},
            {"key": "email", "label_pt": "E-mail", "label_en": "Email", "getter": lambda item, lang: item.email},
            {"key": "position", "label_pt": "Cargo", "label_en": "Position", "getter": lambda item, lang: _localized(item.positions.first(), "name", lang) if item.positions.exists() else ""},
            {"key": "roles", "label_pt": "Funções", "label_en": "Roles", "getter": lambda item, lang: ", ".join(item.roles) if item.roles else ""},
            {"key": "room", "label_pt": "Sala", "label_en": "Room", "getter": lambda item, lang: item.room.name if item.room else ""},
        ],
        "get_data": lambda: User.objects.filter(is_approved=True).select_related("room").prefetch_related("positions").order_by("name"),
        "supports_room_grouping": True,
        "supports_section_grouping": False,
    },
    "equipment": {
        "label_pt": "Equipamentos",
        "label_en": "Equipment",
        "columns": [
            {"key": "custom_id", "label_pt": "ID Interno", "label_en": "Internal ID", "getter": lambda item, lang: item.custom_id},
            {"key": "name", "label_pt": "Nome", "label_en": "Name", "getter": lambda item, lang: item.name},
            {"key": "observation", "label_pt": "Observação", "label_en": "Observation", "getter": lambda item, lang: item.observation or ""},
            {"key": "category", "label_pt": "Categoria", "label_en": "Category", "getter": lambda item, lang: item.identification_category.name if item.identification_category else ""},
            {"key": "state", "label_pt": "Estado", "label_en": "State", "getter": lambda item, lang: item.equipment_state.name if item.equipment_state else ""},
            {"key": "room", "label_pt": "Sala", "label_en": "Room", "getter": lambda item, lang: item.room.name if item.room else ""},
            {"key": "section", "label_pt": "Seção", "label_en": "Section", "getter": lambda item, lang: item.section.name if item.section else ""},
            {"key": "assigned_to", "label_pt": "Responsável", "label_en": "Responsible", "getter": lambda item, lang: item.assigned_to.name if item.assigned_to else ""},
            {"key": "status", "label_pt": "Status", "label_en": "Status", "getter": lambda item, lang: _status_label(item.is_active, lang)},
        ],
        "get_data": lambda: Equipment.objects.select_related("assigned_to", "room", "section", "identification_category", "equipment_state").all().order_by("order", "name"),
        "supports_room_grouping": True,
        "supports_section_grouping": True,
    },
    "partnerships": {
        "label_pt": "Parcerias",
        "label_en": "Partnerships",
        "columns": [
            {"key": "id", "label_pt": "ID", "label_en": "ID", "getter": lambda item, lang: item.id},
            {"key": "name", "label_pt": "Nome", "label_en": "Name", "getter": lambda item, lang: item.name},
            {"key": "link", "label_pt": "Link", "label_en": "Link", "getter": lambda item, lang: item.link or ""},
            {"key": "status", "label_pt": "Status", "label_en": "Status", "getter": lambda item, lang: _status_label(item.is_active, lang)},
        ],
        "get_data": lambda: Partnership.objects.all().order_by("order", "name"),
        "supports_room_grouping": False,
        "supports_section_grouping": False,
    },
}

VALID_SECTIONS = set(COLUMN_CONFIG.keys())


@login_required
@require_http_methods(["GET"])
def get_report_config(request):
    if not hasattr(request.user, "has_role") or not request.user.has_role("professor"):
        return JsonResponse({"error": "Permissão negada."}, status=403)

    config_data = {}
    for section_key, config in COLUMN_CONFIG.items():
        config_data[section_key] = {
            "label_pt": config["label_pt"],
            "label_en": config["label_en"],
            "columns": [
                {"key": col["key"], "label_pt": col["label_pt"], "label_en": col["label_en"]}
                for col in config["columns"]
            ],
            "supports_room_grouping": config["supports_room_grouping"],
            "supports_section_grouping": config["supports_section_grouping"],
        }

    return JsonResponse({"success": True, "data": config_data})


@login_required
@require_http_methods(["POST"])
def generate_report(request):
    if not hasattr(request.user, "has_role") or not request.user.has_role("professor"):
        return JsonResponse({"error": "Permissão negada."}, status=403)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)

    report_name = data.get("name", "").strip()
    output_format = data.get("format", "pdf").lower()
    sections_data = data.get("sections", {})
    lang = _normalize_lang(data.get("language"))
    report_tz = _resolve_timezone(data.get("timezone"))

    if not report_name:
        return JsonResponse({"error": "O nome do relatório é obrigatório."}, status=400)

    if output_format not in ("pdf", "xlsx"):
        return JsonResponse({"error": "Formato inválido. Use 'pdf' ou 'xlsx'."}, status=400)

    if not sections_data or not isinstance(sections_data, dict):
        return JsonResponse({"error": "Selecione ao menos uma seção."}, status=400)

    valid_sections = {}
    for section_key, section_opts in sections_data.items():
        if section_key not in VALID_SECTIONS:
            continue
        if not isinstance(section_opts, dict):
            continue
        columns = section_opts.get("columns", [])
        if not columns:
            continue
        config = COLUMN_CONFIG[section_key]
        valid_column_keys = {col["key"] for col in config["columns"]}
        filtered_columns = [c for c in columns if c in valid_column_keys]
        if not filtered_columns:
            continue
        valid_sections[section_key] = {
            "columns": filtered_columns,
            "group_by_room": section_opts.get("group_by_room", False) and config["supports_room_grouping"],
            "group_by_section": section_opts.get("group_by_section", False) and config["supports_section_grouping"],
        }

    if not valid_sections:
        return JsonResponse({"error": "Nenhuma seção válida selecionada."}, status=400)

    if output_format == "xlsx":
        return _generate_xlsx(report_name, valid_sections, lang)
    else:
        return _generate_pdf(report_name, valid_sections, lang, report_tz)


def _get_column_info(section_key, selected_columns, lang):
    config = COLUMN_CONFIG[section_key]
    column_map = {col["key"]: col for col in config["columns"]}
    headers = []
    getters = []
    for col_key in selected_columns:
        if col_key in column_map:
            headers.append(column_map[col_key][f"label_{lang}"])
            getters.append(column_map[col_key]["getter"])
    return headers, getters


def _group_items_by_room_section(items, section_key, group_by_room, group_by_section, lang):
    if not group_by_room:
        return [{"label": None, "items": list(items)}]
    
    grouped = {}
    no_room_items = []
    
    for item in items:
        room = getattr(item, "room", None)
        if not room:
            no_room_items.append(item)
            continue
        
        if group_by_section and section_key == "equipment":
            section = getattr(item, "section", None)
            if section:
                key = (room.id, section.id)
                label = f"{room.name} - {section.name}"
            else:
                key = (room.id, None)
                label = room.name
        else:
            key = (room.id, None)
            label = room.name
        
        if key not in grouped:
            grouped[key] = {"label": label, "items": [], "room_order": room.order, "section_order": getattr(getattr(item, "section", None), "order", 0) if group_by_section else 0}
        grouped[key]["items"].append(item)
    
    result = sorted(grouped.values(), key=lambda x: (x["room_order"], x["section_order"]))
    if no_room_items:
        result.append({"label": UI_STRINGS[lang]["no_room"], "items": no_room_items})
    
    return result


def _generate_xlsx(report_name, sections_data, lang):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D7A73", end_color="2D7A73", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0E8E6"),
        right=Side(style="thin", color="D0E8E6"),
        top=Side(style="thin", color="D0E8E6"),
        bottom=Side(style="thin", color="D0E8E6"),
    )
    group_font = Font(bold=True, color="2D7A73", size=12)

    for section_key, section_opts in sections_data.items():
        config = COLUMN_CONFIG[section_key]
        selected_columns = section_opts["columns"]
        group_by_room = section_opts["group_by_room"]
        group_by_section = section_opts["group_by_section"]
        
        headers, getters = _get_column_info(section_key, selected_columns, lang)
        ws = wb.create_sheet(title=config[f"label_{lang}"][:31])

        items = config["get_data"]()
        groups = _group_items_by_room_section(items, section_key, group_by_room, group_by_section, lang)
        
        current_row = 1
        
        for group in groups:
            if group["label"]:
                cell = ws.cell(row=current_row, column=1, value=group["label"])
                cell.font = group_font
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                current_row += 1
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1
            
            for item in group["items"]:
                for col_idx, getter in enumerate(getters, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=getter(item, lang))
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                current_row += 1
            
            current_row += 1

        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = report_name.replace(" ", "_")
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
    return response


def _generate_pdf(report_name, sections_data, lang, tz):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#2D7A73")
    elements.append(Paragraph(report_name, title_style))

    subtitle_style = styles["Normal"].clone("subtitle")
    subtitle_style.textColor = colors.HexColor("#6B9591")
    subtitle_style.fontSize = 10
    generated_at = timezone.now().astimezone(tz).strftime(UI_STRINGS[lang]["datetime_format"])
    elements.append(Paragraph(f"{UI_STRINGS[lang]['generated_at']} {generated_at}", subtitle_style))
    elements.append(Spacer(1, 0.8 * cm))

    header_color = colors.HexColor("#2D7A73")
    header_text_color = colors.white
    border_color = colors.HexColor("#D0E8E6")
    alt_row_color = colors.HexColor("#F5F9F8")

    cell_style = styles["Normal"].clone("cell")
    cell_style.fontSize = 8
    cell_style.leading = 10

    header_cell_style = styles["Normal"].clone("header_cell")
    header_cell_style.fontSize = 8
    header_cell_style.leading = 10
    header_cell_style.textColor = header_text_color
    header_cell_style.fontName = "Helvetica-Bold"

    group_title_style = styles["Normal"].clone("group_title")
    group_title_style.fontSize = 10
    group_title_style.textColor = colors.HexColor("#2D7A73")
    group_title_style.fontName = "Helvetica-Bold"

    for section_key, section_opts in sections_data.items():
        config = COLUMN_CONFIG[section_key]
        selected_columns = section_opts["columns"]
        group_by_room = section_opts["group_by_room"]
        group_by_section = section_opts["group_by_section"]
        
        headers, getters = _get_column_info(section_key, selected_columns, lang)

        section_title_style = styles["Heading2"].clone(f"section_{section_key}")
        section_title_style.textColor = colors.HexColor("#2D7A73")
        section_title_style.fontSize = 14
        elements.append(Paragraph(config[f"label_{lang}"], section_title_style))
        elements.append(Spacer(1, 0.3 * cm))

        items = list(config["get_data"]())

        if not items:
            empty_style = styles["Normal"].clone(f"empty_{section_key}")
            empty_style.textColor = colors.HexColor("#6B9591")
            elements.append(Paragraph(UI_STRINGS[lang]["no_records"], empty_style))
            elements.append(Spacer(1, 0.6 * cm))
            continue

        groups = _group_items_by_room_section(items, section_key, group_by_room, group_by_section, lang)

        for group in groups:
            if group["label"]:
                elements.append(Paragraph(group["label"], group_title_style))
                elements.append(Spacer(1, 0.2 * cm))

            table_data = [[Paragraph(h, header_cell_style) for h in headers]]
            for item in group["items"]:
                row = [getter(item, lang) for getter in getters]
                table_data.append([Paragraph(str(v), cell_style) for v in row])

            num_cols = len(headers)
            available_width = landscape(A4)[0] - 3 * cm
            col_width = available_width / num_cols

            table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
            style_commands = [
                ("BACKGROUND", (0, 0), (-1, 0), header_color),
                ("TEXTCOLOR", (0, 0), (-1, 0), header_text_color),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, border_color),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]

            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    style_commands.append(("BACKGROUND", (0, i), (-1, i), alt_row_color))

            table.setStyle(TableStyle(style_commands))
            elements.append(table)
            elements.append(Spacer(1, 0.5 * cm))

        elements.append(Spacer(1, 0.3 * cm))

    doc.build(elements)
    buffer.seek(0)

    safe_name = report_name.replace(" ", "_")
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{safe_name}.pdf"'
    return response
